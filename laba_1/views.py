from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.http import JsonResponse, HttpResponseForbidden
from django.db.models import Q
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
import os
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .forms import (
    CompetitionForm, AthleteSearchForm, CompetitionSelectionForm, 
    AthleteResultForm, AthleteForm, UserRegistrationForm, UserLoginForm,
    UserProfileForm, PasswordChangeForm
)
from .models import (
    Athlete, Result, Competition, Distance, UserProfile, 
    EmailConfirmation, ActivityLog, UserGroup
)
from laba_1.core.utils import get_points_by_place, parse_time_to_seconds, calculate_rank_for_result
import uuid
from datetime import timedelta
from django.utils import timezone
from django.core.cache import cache
import hashlib


def competition_registration(request):
    form = CompetitionForm()
    submitted_data = None

    if request.method == 'POST':
        form = CompetitionForm(request.POST)
        if form.is_valid():
            submitted_data = {
                'full_name': form.cleaned_data['full_name'],
                'email': form.cleaned_data['email'],
                'region': form.cleaned_data['region'],
                'agree_pd': form.cleaned_data['agree_pd'],
            }


    context = {
        'form': form,
        'submitted_data': submitted_data,
    }

    return render(request, 'competition_registration.html', context)



# Задание 8 жестко да
def athlete_points_calculation(request):

    search_form = AthleteSearchForm()
    selection_form = CompetitionSelectionForm()


    athlete = None
    results = []
    total_points = 0
    error_message = None
    best_results_by_distance = {}
    all_calculated_results = []
    distance_best_results = {}


    if request.method == 'POST':
        search_form = AthleteSearchForm(request.POST)
        selection_form = CompetitionSelectionForm(request.POST)

        if search_form.is_valid() and selection_form.is_valid():


            athlete_name = search_form.cleaned_data['athlete_name']
            target_competition = selection_form.cleaned_data['target_competition']
            history_competitions = selection_form.cleaned_data['history_competitions']


            athletes = Athlete.objects.filter(full_name__icontains=athlete_name).order_by('full_name')

            if athletes.exists():
                athlete = athletes.first()

                all_results = Result.objects.filter(
                    athlete=athlete
                ).select_related('competition', 'distance').order_by(
                    'competition__date', 'distance__distance_meters'
                )


                if history_competitions.exists():
                    results = all_results.filter(competition__in=history_competitions)

                target_results_by_distance = {}
                for r in Result.objects.filter(
                    competition=target_competition
                ).select_related('distance').filter(
                    result_time__isnull=False
                ):
                    dist_key = str(r.distance)
                    if dist_key not in target_results_by_distance:
                        target_results_by_distance[dist_key] = []
                    time_sec = parse_time_to_seconds(r.result_time)
                    if time_sec is not None:
                        target_results_by_distance[dist_key].append({
                            'time': time_sec,
                            'place': r.place,
                            'athlete': r.athlete.full_name
                        })


                for dist_key in target_results_by_distance:
                    target_results_by_distance[dist_key].sort(key=lambda x: x['time'])

                all_calculated_results = []
                for r in results:
                    dist_name = str(r.distance)

                    if not r.result_time:
                        continue

                    athlete_time_sec = parse_time_to_seconds(r.result_time)
                    if athlete_time_sec is None:
                        continue

                    is_200_underwater = (
                        'Подводное плавание' in dist_name and
                        '200' in dist_name
                    )

                    predicted_place = None
                    if dist_name in target_results_by_distance:
                        target_times = target_results_by_distance[dist_name]

                        faster_count = sum(1 for t in target_times if t['time'] < athlete_time_sec)
                        predicted_place = faster_count + 1


                    if is_200_underwater:
                        calculated_points = 0
                    elif predicted_place is not None:
                        calculated_points = get_points_by_place(predicted_place)
                    else:
                        calculated_points = 0

                    calculated_rank = calculate_rank_for_result(r.distance, r.result_time)

                    all_calculated_results.append({
                        'result': r,
                        'points': calculated_points,
                        'predicted_place': predicted_place,
                        'athlete_time_sec': athlete_time_sec,
                        'calculated_rank': calculated_rank
                    })

                    if dist_name not in distance_best_results:
                        distance_best_results[dist_name] = {
                            'result': r,
                            'points': calculated_points,
                            'predicted_place': predicted_place,
                            'athlete_time_sec': athlete_time_sec
                        }
                    else:

                        if calculated_points > distance_best_results[dist_name]['points']:
                            distance_best_results[dist_name] = {
                                'result': r,
                                'points': calculated_points,
                                'predicted_place': predicted_place,
                                'athlete_time_sec': athlete_time_sec
                            }

                sorted_by_points = sorted(
                    distance_best_results.values(),
                    key=lambda x: x['points'],
                    reverse=True
                )[:3]

                total_points = sum(item['points'] for item in sorted_by_points)

                best_results_by_distance = distance_best_results


            else:
                error_message = f"Спортсмен с именем '{athlete_name}' не найден в базе"
        else:
            error_message = "Проверьте правильность заполнения формы"

    context = {
        'search_form': search_form,
        'selection_form': selection_form,
        'athlete': athlete,
        'results': results,
        'total_points': total_points,
        'error_message': error_message,
        'best_results_by_distance': best_results_by_distance,
        'all_calculated_results': all_calculated_results,
    }

    return render(request, 'athlete_points.html', context)



@login_required(login_url='login')
def add_member(request):

    form = AthleteResultForm()

    if request.method == 'POST':
        form = AthleteResultForm(request.POST)

        if form.is_valid():

            athlete = Athlete.objects.create(
                full_name=form.cleaned_data['athlete_full_name'],
                birth_year=form.cleaned_data['athlete_birth_year'],
                team=form.cleaned_data.get('athlete_team', '')
            )


            result = Result.objects.create(
                athlete=athlete,
                competition=form.cleaned_data['competition'],
                distance=form.cleaned_data['distance'],
                result_time=form.cleaned_data['result_time'],
            )
            cache.clear()

            log_activity(
                request.user,
                'add_athlete',
                f'Добавлен спортсмен: {athlete.full_name}',
                request=request
            )

            messages.success(
                request,
                f'Спортсмен "{athlete.full_name}" успешно добавлен! '
            )

            return redirect('add_member')

    context = {
        'form': form,
        'page_title': 'Добавить спортсмена'
    }

    return render(request, 'add_member.html', context)

@login_required(login_url='login')
def athlete_detail(request, pk):
    athlete = get_object_or_404(Athlete, pk=pk, removed=False)
    results = Result.objects.filter(athlete=athlete).select_related('competition', 'distance')
    return render(request, 'athlete_detail.html', {'athlete': athlete, 'results': results})


@login_required(login_url='login')
def athlete_edit(request, pk):
    athlete = get_object_or_404(Athlete, pk=pk, removed=False)

    if request.method == 'POST':
        form = AthleteForm(request.POST, instance=athlete)
        if form.is_valid():
            form.save()

            cache.clear()
            log_activity(
                request.user,
                'edit_athlete',
                f'Редактирован спортсмен: {athlete.full_name}',
                request=request
            )
            
            messages.success(request, f'Данные спортсмена "{athlete.full_name}" успешно обновлены!')
            return redirect('athlete_detail', pk=athlete.pk)
    else:
        form = AthleteForm(instance=athlete)

    return render(request, 'athlete_edit.html', {'form': form, 'athlete': athlete, 'page_title': 'Редактирование'})


@login_required(login_url='login')
def athlete_delete(request, pk):
    athlete = get_object_or_404(Athlete, pk=pk)
    athlete_name = athlete.full_name
    athlete.removed = True
    athlete.save()

    cache.clear()
    
    log_activity(
        request.user,
        'delete_athlete',
        f'Удален спортсмен: {athlete_name}',
        request=request
    )
    
    messages.success(request, f'Спортсмен "{athlete_name}" успешно удален!')
    return redirect('athlete_list')



@login_required(login_url='login')
def export_excel_ajax(request):

    athletes = Athlete.objects.filter(removed=False).order_by('full_name')

    wb = Workbook()
    ws = wb.active
    ws.title = "Спортсмены"

    headers = ['№', 'ФИО', 'Год рождения', 'Команда', 'Соревнование', 'Дистанция', 'Результат']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = cell.font.copy(bold=True)


    row_num = 2
    for athlete in athletes:
        results = Result.objects.filter(athlete=athlete).select_related('competition', 'distance')

        if results.exists():
            for result in results:
                ws.cell(row=row_num, column=1, value=row_num - 1)
                ws.cell(row=row_num, column=2, value=athlete.full_name)
                ws.cell(row=row_num, column=3, value=athlete.birth_year)
                ws.cell(row=row_num, column=4, value=athlete.team or '')
                ws.cell(row=row_num, column=5, value=str(result.competition))
                ws.cell(row=row_num, column=6, value=str(result.distance))
                ws.cell(row=row_num, column=7, value=result.result_time or '')
                row_num += 1
        else:
            ws.cell(row=row_num, column=1, value=row_num - 1)
            ws.cell(row=row_num, column=2, value=athlete.full_name)
            ws.cell(row=row_num, column=3, value=athlete.birth_year)
            ws.cell(row=row_num, column=4, value=athlete.team or '')
            row_num += 1


    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width


    filename = 'athletes_export.xlsx'
    file_path = os.path.join(settings.MEDIA_ROOT, 'excel_exports', filename)

    wb.save(file_path)

    log_activity(
        request.user,
        'export',
        f'Экспортировано {row_num - 2} записей в Excel',
        request=request
    )

    file_url = settings.MEDIA_URL + 'excel_exports/' + filename

    return JsonResponse({
        'success': True,
        'file_url': file_url,
        'filename': filename,
        'message': f'Экспортировано {row_num - 2} записей'
    })


@require_http_methods(["POST"])
@login_required(login_url='login')
def athlete_delete_ajax(request, pk):
    try:
        athlete = get_object_or_404(Athlete, pk=pk)
        athlete_name = athlete.full_name
        athlete.removed = True
        athlete.save()

        log_activity(
            request.user,
            'delete_athlete',
            f'Удален спортсмен (AJAX): {athlete_name}',
            request=request
        )

        return JsonResponse({
            'success': True,
            'message': f'Спортсмен "{athlete_name}" успешно удален!',
            'deleted_id': pk
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Ошибка при удалении: {str(e)}'
        })



def athlete_list(request):

    search_query = request.GET.get('search', '').strip()
    team_filter = request.GET.get('team', '').strip()
    page_number = request.GET.get('page', 1)
    
    # Создаем уникальный ключ кэша на основе всех параметров
    cache_key_data = f'athlete_list_{search_query}_{team_filter}_{page_number}'
    cache_key = f'athlete_list_{hashlib.md5(cache_key_data.encode()).hexdigest()}'
    
    # Пробуем получить данные из кэша
    cached_context = cache.get(cache_key)
    
    if cached_context:
        
        return render(request, 'athlete_list.html', cached_context)
    
    

    athletes = Athlete.objects.filter(removed=False).select_related()
    
    # Поиск
    if search_query:
        athletes = athletes.filter(
            Q(full_name__icontains=search_query) | Q(team__icontains=search_query)
        )
    
    # Фильтрация по команде
    if team_filter:
        athletes = athletes.filter(team=team_filter)
    
    # AJAX запрос
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        data = {
            'success': True,
            'count': athletes.count(),
            'athletes': []
        }
        for i, athlete in enumerate(athletes, start=1):
            data['athletes'].append({
                'id': athlete.pk,
                'row_number': i,
                'full_name': athlete.full_name,
                'birth_year': athlete.birth_year,
                'team': athlete.team,
                'detail_url': f'/athletes/{athlete.pk}/'
            })
        return JsonResponse(data)
    
    # Пагинация
    paginator = Paginator(athletes, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Получить список всех команд для фильтра
    teams = Athlete.objects.filter(removed=False).values_list('team', flat=True).distinct()
    teams = [t for t in teams if t]
    
    context = {
        'athletes': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'team_filter': team_filter,
        'teams': sorted(teams),
        'total_count': athletes.count(),
        'page_title': 'Список спортсменов'
    }


    cache.set(cache_key, context, settings.CACHE_TTL)
    print(f"[CACHE SET] Сохранено в кэш на {settings.CACHE_TTL} секунд")
    
    return render(request, 'athlete_list.html', context)


def get_client_ip(request):
    """Получить IP адрес клиента"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_activity(user, action, description="", request=None):
    """Логировать действие пользователя"""
    ip = get_client_ip(request) if request else None
    user_agent = request.META.get('HTTP_USER_AGENT', '') if request else ''
    
    ActivityLog.objects.create(
        user=user,
        action=action,
        description=description,
        ip_address=ip,
        user_agent=user_agent
    )


def register(request):
    """Регистрация нового пользователя"""
    if request.user.is_authenticated:
        return redirect('athlete_list')
    
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            # Создать пользователя
            user = User.objects.create_user(
                username=form.cleaned_data['username'],
                email=form.cleaned_data['email'],
                password=form.cleaned_data['password'],
                first_name=form.cleaned_data.get('first_name', ''),
                last_name=form.cleaned_data.get('last_name', '')
            )
            
            # Получить группу "athlete" по умолчанию
            try:
                athlete_group = UserGroup.objects.get(name='athlete')
            except UserGroup.DoesNotExist:
                athlete_group = UserGroup.objects.create(
                    name='athlete',
                    description='Спортсмен (редактирование)'
                )
            
            # Создать профиль пользователя
            profile = UserProfile.objects.create(
                user=user,
                group=athlete_group
            )
            
            # Создать токен подтверждения email
            token = uuid.uuid4()
            EmailConfirmation.objects.create(
                user=user,
                email=user.email,
                token=str(token),
                expires_at=timezone.now() + timedelta(hours=24)
            )
            
            # СОЗДАЕМ ПЕРЕМЕННУЮ confirmation_url (ВОТ ЧЕГО НЕ ХВАТАЛО)
            confirmation_url = request.build_absolute_uri(f'/confirm-email/{token}/')
            
            # ВЫВОДИМ ССЫЛКУ В КОНСОЛЬ
            print("\n" + "="*70)
            print("📧 ДЛЯ ПОДТВЕРЖДЕНИЯ EMAIL ПЕРЕЙДИТЕ ПО ССЫЛКЕ:")
            print(confirmation_url)
            print("="*70 + "\n")
            
            # Отправить email с подтверждением
            subject = 'Подтвердите вашу почту'
            message = f'Подтвердите вашу почту: {confirmation_url}'
            
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [user.email],
                fail_silently=True,
            )
            
            messages.success(
                request, 
                'Регистрация успешна! Проверьте почту для подтверждения email.'
            )
            return redirect('login')
    else:
        form = UserRegistrationForm()
    
    context = {'form': form, 'page_title': 'Регистрация'}
    return render(request, 'auth/register.html', context)

def login_view(request):
    # """Вход в аккаунт"""
    # if request.user.is_authenticated:
    #     return redirect('athlete_list')
    
    if request.method == 'POST':
        form = UserLoginForm(request.POST)
        if form.is_valid():
            user = authenticate(
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password']
            )
            
            if user is not None:
                login(request, user)
                log_activity(user, 'login', request=request)
                
                if form.cleaned_data.get('remember_me'):
                    request.session.set_expiry(timedelta(days=30))
                
                messages.success(request, f'Добро пожаловать, {user.first_name or user.username}!')
                return redirect('athlete_list')
            else:
                messages.error(request, 'Неверное имя пользователя или пароль')
    else:
        form = UserLoginForm()
    
    context = {'form': form, 'page_title': 'Вход'}
    return render(request, 'auth/login.html', context)


@login_required(login_url='login')
def logout_view(request):
    """Выход из аккаунта"""
    log_activity(request.user, 'logout', request=request)
    logout(request)
    messages.success(request, 'Вы вышли из аккаунта')
    return redirect('login')


@login_required(login_url='login')
def profile_view(request):
    """Профиль пользователя"""
    profile = request.user.profile
    
    context = {
        'profile': profile,
        'user': request.user,
        'page_title': 'Мой профиль'
    }
    return render(request, 'auth/profile.html', context)


@login_required(login_url='login')
def profile_edit(request):
    """Редактирование профиля"""
    profile = request.user.profile
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        
        # Обновить данные User
        request.user.first_name = request.POST.get('first_name', '')
        request.user.last_name = request.POST.get('last_name', '')
        request.user.email = request.POST.get('email', '')
        
        if form.is_valid():
            form.save()
            request.user.save()
            
            log_activity(
                request.user, 
                'profile_update', 
                'Обновлены данные профиля',
                request=request
            )
            
            messages.success(request, 'Профиль успешно обновлен!')
            return redirect('profile')
    else:
        form = UserProfileForm(instance=profile)
        # Заполнить поля User
        form.initial['first_name'] = request.user.first_name
        form.initial['last_name'] = request.user.last_name
        form.initial['email'] = request.user.email
    
    context = {
        'form': form,
        'profile': profile,
        'page_title': 'Редактирование профиля'
    }
    return render(request, 'auth/profile_edit.html', context)


@login_required(login_url='login')
def password_change(request):
    """Изменение пароля"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.POST)
        
        # Проверить текущий пароль
        if not request.user.check_password(form.data.get('old_password')):
            messages.error(request, 'Текущий пароль неверный')
            form = PasswordChangeForm()
        elif form.is_valid():
            request.user.set_password(form.cleaned_data['new_password'])
            request.user.save()
            
            log_activity(
                request.user,
                'other',
                'Изменен пароль',
                request=request
            )
            
            messages.success(request, 'Пароль успешно изменен!')
            return redirect('profile')
    else:
        form = PasswordChangeForm()
    
    context = {
        'form': form,
        'page_title': 'Изменение пароля'
    }
    return render(request, 'auth/password_change.html', context)


def confirm_email(request, token):
    """Подтверждение email"""
    try:
        confirmation = EmailConfirmation.objects.get(token=token)
        
        if confirmation.is_confirmed:
            messages.info(request, 'Email уже был подтвержден ранее')
        elif confirmation.expires_at and confirmation.expires_at < timezone.now():
            messages.error(request, 'Ссылка подтверждения истекла')
        else:
            confirmation.is_confirmed = True
            confirmation.confirmed_at = timezone.now()
            confirmation.save()
            
            # Обновить профиль
            profile = confirmation.user.profile
            profile.email_verified = True
            profile.save()
            
            messages.success(request, 'Email успешно подтвержден!')
    except EmailConfirmation.DoesNotExist:
        messages.error(request, 'Некорректная ссылка подтверждения')
    
    return redirect('login')


@login_required(login_url='login')
def activity_log_view(request):
    """История активности пользователя"""
    activities = ActivityLog.objects.filter(user=request.user).order_by('-created_at')
    
    # Пагинация
    paginator = Paginator(activities, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'activities': page_obj,
        'page_obj': page_obj,
        'page_title': 'История активности'
    }
    return render(request, 'auth/activity_log.html', context)


def competition_list(request):
    """Список соревнований с фильтрацией и поиском"""
    competitions = Competition.objects.all().order_by('-date')
    
    # Поиск по названию
    search_query = request.GET.get('q', '')
    if search_query:
        competitions = competitions.filter(
            Q(name__icontains=search_query) |
            Q(city__icontains=search_query) |
            Q(description__icontains=search_query)
        )
        log_activity(request.user, 'other', f'Поиск соревнований: {search_query}', request=request) if request.user.is_authenticated else None
    
    

    
    # Пагинация (20 на странице)
    paginator = Paginator(competitions, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    
    context = {
        'competitions': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'page_title': 'Список соревнований',
        'total_count': Competition.objects.count(),
    }
    return render(request, 'competition_list.html', context)


